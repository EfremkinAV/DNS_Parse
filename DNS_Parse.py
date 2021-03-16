from selenium import webdriver
import re
from selenium.common.exceptions import NoSuchElementException
import sys
import sqlite3
import traceback
from datetime import datetime
from bs4 import BeautifulSoup
import requests
import io
from webdriver_manager.chrome import ChromeDriverManager
import urllib
from urllib.request import urlretrieve
import zipfile
import xlrd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import time
import os


def save_file(city):
    city = "price-" + city + ".zip"
    url = 'https://www.dns-shop.ru/files/price/'+city
    urllib.request.urlretrieve(url, city)
    print(city)
    file_zip = zipfile.ZipFile(city, "r")
    file_zip.extractall('')
    file_name = str(''.join(file_zip.namelist()))
    file_size = file_zip.getinfo(file_name).file_size
    print("Распакован файл: ", file_name)
    print("Размер файла: ", file_size / 1000000, "Mb")
    print("Дата создания файла: ", time.ctime(os.path.getctime(city)))
    file_zip.close()

#save_file("tomsk")#input("Введите город:"))


def fill_db(city):
    start_time = time.time()
    sqlite_connection = sqlite3.connect("DNS_PARSE.db")
    con = sqlite_connection.cursor()

    #city = "tomsk"  # input("Введите город: ")
    file_name = "price-" + city + ".xls"
    book = xlrd.open_workbook(file_name)
    sheets_num = book.nsheets
    for i in range(1, sheets_num - 1):
        rows = book.sheet_by_index(i)
        for r in range(13, rows.nrows):

            kod = rows.cell_value(r, 0)
            prod = rows.cell_value(r, 1)
            M1 = rows.cell_value(r, 2)
            M2 = rows.cell_value(r, 3)
            M3 = rows.cell_value(r, 4)
            M4 = rows.cell_value(r, 5)
            M5 = rows.cell_value(r, 6)
            M6 = rows.cell_value(r, 7)
            M7 = rows.cell_value(r, 8)
            M8 = rows.cell_value(r, 9)
            M9 = rows.cell_value(r, 10)
            M10 = rows.cell_value(r, 11)
            M11 = rows.cell_value(r, 12)
            price = rows.cell_value(r, 13)
            parse_date = datetime.now().date() #Добыть дату файла

            con.execute("INSERT INTO product (kod,prod,M1,M2,M3,M4,M5,M6,M7,M8,M9,M10,M11,price,date)"
                        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (kod, prod, M1, M2, M3, M4, M5, M6, M7, M8, M9, M10, M11, price, parse_date))
            sqlite_connection.commit()

    con.close()
    sqlite_connection.close()


    print("Таблица распарсилась за %s секунд" % (time.time() - start_time))
    print("---КОНЕЦ---")

#fill_db("tomsk")

def db_delete_table(table_name):
    sqlite_connection = sqlite3.connect('DNS_PARSE.db')
    con = sqlite_connection.cursor()
    con.execute("DELETE from " + table_name)
    sqlite_connection.commit()
    con.close()
    sqlite_connection.close()
    print("ТАБЛИЦА", table_name, "ОЧИЩЕНА!")

db_delete_table("product")
